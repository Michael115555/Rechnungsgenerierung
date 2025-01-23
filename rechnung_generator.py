import os
from datetime import datetime
from openpyxl import load_workbook
from qrbill import QRBill
from wand.image import Image as WandImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_LEFT, TA_CENTER


def svg_to_png(svg_path, png_path):
    try:
        with WandImage(filename=svg_path) as img:
            img.format = 'png'
            img.save(filename=png_path)
    except Exception as e:
        print(f"Fehler bei der Konvertierung der SVG zu PNG: {e}")


def lade_kundendaten_und_positionen(dateiname):
    wb = load_workbook(dateiname)
    
    firmendaten = {
        'Name': 'Michael Vogel',
        'Strasse': 'Ihre Straße',
        'PLZ': '1234',
        'Ort': 'Ihre Stadt',
        'IBAN': 'CH9300762011623852957'  # Test-IBAN, die das Format erfüllt
    }

    ws_kunde = wb['Kunden']
    kundendaten_liste = []
    for row in ws_kunde.iter_rows(min_row=2, values_only=True):
        if all(row[:4]):
            kundendaten_liste.append({
                'Kundennummer': row[0],
                'Firma': row[1],
                'Kontakt': row[2],
                'Adresse': row[3],
                'PLZ': str(row[4]),
                'Ort': row[5],
                'IBAN': row[6],
                'Email': row[7]
            })

    ws_positionen = wb['Positionen']
    positionen = []
    for row in ws_positionen.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2] and row[3]:
            try:
                menge = float(str(row[2]))
                preis_pro_stunde = float(str(row[3]))
                positionspreis = menge * preis_pro_stunde
                
                datum = row[5]
                if isinstance(datum, datetime):
                    positionen.append({
                        'Kundennummer': row[0],
                        'Bezeichnung': row[1],
                        'Menge': menge,
                        'Preis/Stunde': preis_pro_stunde,
                        'Positionspreis': positionspreis,
                        'Datum': datum
                    })
            except Exception as e:
                continue

    return firmendaten, kundendaten_liste, positionen


def erstelle_rechnung_pdf(firmendaten, kundendaten, positionen, output_path, rechnungsdatum=None):
    if rechnungsdatum is None:
        rechnungsdatum = datetime.today().strftime('%Y%m%d')
    
    kunden_positionen = [p for p in positionen 
                        if p['Kundennummer'] == kundendaten['Kundennummer'] 
                        and p['Datum'].strftime('%Y%m%d') == rechnungsdatum]
    
    if not kunden_positionen:
        return
            
    gesamtbetrag = sum(p['Positionspreis'] for p in kunden_positionen)
    
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='RightAlign',
        parent=styles['Normal'],
        alignment=TA_RIGHT
    ))
    
    absender_style = ParagraphStyle(
        'Absender',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.gray
    )
    
    story = []
    
    story.append(Paragraph(f"{firmendaten['Name']} · {firmendaten['Strasse']} · {firmendaten['PLZ']} {firmendaten['Ort']}", absender_style))
    story.append(Spacer(1, 20))
    
    story.append(Paragraph(f"{kundendaten['Firma']}", styles['Normal']))
    if kundendaten['Kontakt']:
        story.append(Paragraph(f"{kundendaten['Kontakt']}", styles['Normal']))
    story.append(Paragraph(f"{kundendaten['Adresse']}", styles['Normal']))
    story.append(Paragraph(f"{kundendaten['PLZ']} {kundendaten['Ort']}", styles['Normal']))
    story.append(Spacer(1, 40))
    
    positions_datum = kunden_positionen[0]['Datum']
    rechnungsdatum_formatiert = positions_datum.strftime('%d.%m.%Y')
    story.append(Paragraph(f"Datum: {rechnungsdatum_formatiert}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    rechnungsnummer = f"{kundendaten['Kundennummer']}-{positions_datum.strftime('%Y%m%d')}"
    story.append(Paragraph(f"RECHNUNG Nr. {rechnungsnummer}", styles['Heading1']))
    story.append(Spacer(1, 20))
    
    story.append(Paragraph(f"Guten Tag Herr {kundendaten['Kontakt']},", styles['Normal']))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Vielen Dank für den Auftrag. Ich erlaube mir, folgende Leistungen in Rechnung zu stellen.", styles['Normal']))
    story.append(Spacer(1, 20))
    
    data = [['Pos.', 'Bezeichnung', 'Menge', 'Preis/Std', 'Total']]
    for idx, pos in enumerate(kunden_positionen, 1):
        data.append([
            str(idx),
            pos['Bezeichnung'],
            f"{pos['Menge']:.2f}",
            f"CHF {pos['Preis/Stunde']:.2f}",
            f"CHF {pos['Positionspreis']:.2f}"
        ])
    data.append(['', '', '', 'Total:', f"CHF {gesamtbetrag:.2f}"])
    
    table = Table(data, colWidths=[30, 200, 60, 70, 70])
    table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (-2, -1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (-2, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('ALIGN', (-2, 0), (-1, -1), 'RIGHT'),
    ]))
    table.hAlign = 'LEFT'
    
    story.append(table)
    story.append(Spacer(1, 40))
    
    story.append(Paragraph("Zahlungsbedingungen:", styles['Heading2']))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Bitte überweisen Sie den Rechnungsbetrag innerhalb von 20 Tagen. (ab Rechnungsdatum) Nach dem Überweisen vom Rechnungsbetrag gilt das Projekt als abgeschlossen. Für neue Projekte stehe ich Ihnen jederzeit zur Verfügung.", styles['Normal']))
    story.append(Spacer(1, 40))
    
    story.append(Paragraph("Freundliche Grüsse", styles['Normal']))
    story.append(Spacer(1, 20))
    
    unterschrift_pfad = os.path.join('C:\\', 'Users', 'Micooo', 'OneDrive', 'Automatische Rechnung', 'img', 'unterschrift.png')
    if os.path.exists(unterschrift_pfad):
        unterschrift = Image(unterschrift_pfad, width=100, height=50)
        unterschrift.hAlign = 'LEFT'
        story.append(unterschrift)

    story.append(Spacer(1, 20))
    story.append(Paragraph("Michael Vogel", styles['Normal']))
    
    story.append(PageBreak())
    
    qr_svg_path = os.path.join(os.path.dirname(output_path), 'temp_qr.svg')
    qr_png_path = os.path.join(os.path.dirname(output_path), 'temp_qr.png')
    
    qr = QRBill(
        account=firmendaten['IBAN'],
        creditor={
            'name': firmendaten['Name'],
            'street': firmendaten['Strasse'],
            'pcode': firmendaten['PLZ'],
            'city': firmendaten['Ort'],
            'country': 'CH'
        },
        debtor={
            'name': kundendaten['Firma'],
            'street': kundendaten['Adresse'],
            'pcode': kundendaten['PLZ'],
            'city': kundendaten['Ort'],
            'country': 'CH'
        },
        amount=f"{gesamtbetrag:.2f}",
        currency='CHF'
    )
    
    qr.as_svg(qr_svg_path)
    svg_to_png(qr_svg_path, qr_png_path)
    
    page_width = A4[0] - 2*40
    qr_height = page_width / 2
    qr_image = Image(qr_png_path, width=page_width, height=qr_height)
    qr_image.hAlign = 'CENTER'
    story.append(qr_image)
    
    doc.build(story)
    
    try:
        os.remove(qr_svg_path)
        os.remove(qr_png_path)
    except:
        pass


if __name__ == "__main__":
    rechnungen_ordner = os.path.join('C:\\', 'Users', 'Micooo', 'OneDrive', 'Automatische Rechnung', 'Rechnungen')
    os.makedirs(rechnungen_ordner, exist_ok=True)
    
    try:
        # Verwende die bereits existierende Beispieldaten.xlsx
        excel_pfad = os.path.join('C:\\', 'Users', 'Micooo', 'OneDrive', 'Automatische Rechnung', 'Beispieldaten.xlsx')
        firmendaten, kundendaten_liste, positionen = lade_kundendaten_und_positionen(excel_pfad)
        
        unique_dates = set(pos['Datum'].strftime('%Y%m%d') for pos in positionen)
        
        for rechnungsdatum in unique_dates:
            for kunde in kundendaten_liste:
                pdf_dateiname = f"Rechnung_{kunde['Kundennummer']}-{rechnungsdatum}.pdf"
                output_path = os.path.join(rechnungen_ordner, pdf_dateiname)
                erstelle_rechnung_pdf(firmendaten, kunde, positionen, output_path, rechnungsdatum)
            
    except Exception as e:
        print(f"Fehler: {e}")
